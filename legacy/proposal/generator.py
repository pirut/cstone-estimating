from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfgen import canvas

DEFAULT_CONFIG_DIR = Path(__file__).resolve().parent.parent / "config"
DEFAULT_MAPPING_PATH = DEFAULT_CONFIG_DIR / "mapping.json"
DEFAULT_COORDS_PATH = DEFAULT_CONFIG_DIR / "coordinates.json"


def generate_proposal_pdf(
    xlsx_path: str | Path,
    template_pdf_path: str | Path,
    output_pdf_path: str | Path | None = None,
    mapping_path: str | Path | None = None,
    coords_path: str | Path | None = None,
) -> str:
    """Generate a filled Cornerstone Proposal PDF from an Excel workbook."""
    mapping_path = Path(mapping_path) if mapping_path else DEFAULT_MAPPING_PATH
    coords_path = Path(coords_path) if coords_path else DEFAULT_COORDS_PATH

    mapping_config = _load_json(mapping_path)
    coords_config = _load_json(coords_path)

    field_values = _build_field_values(Path(xlsx_path), mapping_config)

    missing_value = mapping_config.get("missing_value", "")
    plan_set_date = field_values.get("plan_set_date")
    if plan_set_date and plan_set_date != missing_value:
        field_values["plan_set_date_line"] = (
            f"Estimate based on plan set dated: {plan_set_date}"
        )
    else:
        field_values["plan_set_date_line"] = missing_value

    reader = PdfReader(str(template_pdf_path))
    writer = PdfWriter()

    for page_index, page in enumerate(reader.pages, start=1):
        page_key = f"page_{page_index}"
        page_fields = coords_config.get(page_key)
        if page_fields:
            width = float(page.mediabox.width)
            height = float(page.mediabox.height)
            overlay_stream = _build_overlay(
                (width, height), page_fields, field_values
            )
            overlay_reader = PdfReader(overlay_stream)
            page.merge_page(overlay_reader.pages[0])
        writer.add_page(page)

    if output_pdf_path is None:
        template_path = Path(template_pdf_path)
        output_pdf_path = template_path.with_name(
            f"{template_path.stem}-filled.pdf"
        )
    else:
        output_pdf_path = Path(output_pdf_path)

    output_pdf_path.parent.mkdir(parents=True, exist_ok=True)
    with output_pdf_path.open("wb") as output_file:
        writer.write(output_file)

    return str(output_pdf_path)


def _load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _build_field_values(
    xlsx_path: Path, mapping_config: Dict[str, Any]
) -> Dict[str, str]:
    field_specs = mapping_config.get("fields", {})
    prepared_by_map = mapping_config.get("prepared_by_map", {})
    missing_value = mapping_config.get("missing_value", "")

    raw_values = _read_excel_values(xlsx_path, field_specs)
    formatted: Dict[str, str] = {}

    for field_name, spec in field_specs.items():
        value = raw_values.get(field_name)
        fmt = spec.get("format", "text")
        formatted[field_name] = _format_value(
            value, fmt, prepared_by_map, missing_value
        )

    return formatted


def _read_excel_values(
    xlsx_path: Path, field_specs: Dict[str, Dict[str, str]]
) -> Dict[str, Any]:
    workbook = load_workbook(filename=str(xlsx_path), data_only=True)
    values: Dict[str, Any] = {}

    for field_name, spec in field_specs.items():
        sheet_name = spec.get("sheet")
        cell = spec.get("cell")
        if not sheet_name or not cell:
            values[field_name] = None
            continue
        if sheet_name not in workbook.sheetnames:
            values[field_name] = None
            continue
        values[field_name] = workbook[sheet_name][cell].value

    workbook.close()
    return values


def _format_value(
    value: Any,
    fmt: str,
    prepared_by_map: Dict[str, str],
    missing_value: str,
) -> str:
    if fmt == "currency":
        return _format_currency(value, missing_value)
    if fmt == "date_cover":
        return _format_date_cover(value, missing_value)
    if fmt == "date_plan":
        return _format_date_plan(value, missing_value)
    if fmt == "initials":
        return _format_initials(value, prepared_by_map, missing_value)
    return _format_text(value, missing_value)


def _format_currency(value: Any, missing_value: str) -> str:
    number = _normalize_number(value)
    if number is None:
        return missing_value
    number = number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"${number:,.2f}"


def _format_date_cover(value: Any, missing_value: str) -> str:
    parsed = _parse_date(value)
    if parsed is None:
        return missing_value
    month = parsed.strftime("%B").upper()
    return f"{month} {parsed.day:02d}, {parsed.year}"


def _format_date_plan(value: Any, missing_value: str) -> str:
    parsed = _parse_date(value)
    if parsed is None:
        return missing_value
    month = parsed.strftime("%B")
    return f"{month} {parsed.day}, {parsed.year}"


def _format_initials(
    value: Any, prepared_by_map: Dict[str, str], missing_value: str
) -> str:
    if value is None:
        return missing_value
    initials = str(value).strip()
    if not initials:
        return missing_value
    return prepared_by_map.get(initials, initials)


def _format_text(value: Any, missing_value: str) -> str:
    if value is None:
        return missing_value
    text = str(value).strip()
    return text if text else missing_value


def _normalize_number(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        cleaned = value.replace("$", "").replace(",", "").strip()
        if not cleaned:
            return None
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None
    return None


def _parse_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def _build_overlay(
    page_size: Tuple[float, float],
    page_fields: Dict[str, Dict[str, Any]],
    field_values: Dict[str, str],
) -> BytesIO:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=page_size)

    for field_name, spec in page_fields.items():
        value = field_values.get(field_name)
        if value is None or value == "":
            continue
        x = float(spec.get("x", 0))
        y = float(spec.get("y", 0))
        font = spec.get("font", "Helvetica")
        size = spec.get("size", 10)
        align = spec.get("align", "left")
        max_width = spec.get("max_width")
        if max_width is not None:
            max_width = float(max_width)
        min_size = spec.get("min_size", 8)

        text, final_size = _fit_text(
            str(value), font, size, max_width, min_size
        )
        pdf.setFont(font, final_size)
        if align == "right":
            pdf.drawRightString(x, y, text)
        elif align == "center":
            pdf.drawCentredString(x, y, text)
        else:
            pdf.drawString(x, y, text)

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return buffer


def _fit_text(
    value: str,
    font: str,
    size: float,
    max_width: Optional[float],
    min_size: float,
) -> Tuple[str, float]:
    if not value:
        return value, size
    if not max_width:
        return value, size

    current = float(size)
    while current >= min_size:
        if pdfmetrics.stringWidth(value, font, current) <= max_width:
            return value, current
        current -= 0.5

    ellipsis = "..."
    if pdfmetrics.stringWidth(ellipsis, font, min_size) > max_width:
        return ellipsis, min_size

    for length in range(len(value), 0, -1):
        candidate = value[:length].rstrip()
        candidate = f"{candidate}{ellipsis}"
        if pdfmetrics.stringWidth(candidate, font, min_size) <= max_width:
            return candidate, min_size

    return ellipsis, min_size
